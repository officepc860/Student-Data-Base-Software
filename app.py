import uuid
from flask import Flask, render_template, request, redirect, url_for, session, flash
from werkzeug.security import generate_password_hash, check_password_hash
from flask_migrate import Migrate
from models import db, User, Student, Payment, NewPayment


app = Flask(__name__)
app.secret_key = 'your_secret_key_here'


app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:1234@localhost/studentdb'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False


db.init_app(app)
migrate = Migrate(app, db)

def join_user_student(query):
    return query.join(Student, User.id == Student.user_id)

with app.app_context():
    db.create_all()




@app.route('/')
def home():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    else:
        return redirect(url_for('login'))


@app.route('/dashboard')
def dashboard():
    if 'user' not in session:
        return redirect(url_for('login'))

    user = session['user']
    user_id = user.get('id')

    selected_batch = request.args.get('batch', '')

    query = Student.query.filter_by(user_id=user_id)
    if selected_batch:
        query = query.filter_by(batch=selected_batch)

    user_students = query.all()
    all_students = Student.query.filter_by(user_id=user_id).all()
    batches = sorted(set(s.batch for s in all_students if s.batch))

    return render_template(
        'dashboard.html',
        students=user_students,
        user=user,
        batches=batches,
        selected_batch=selected_batch
    )
@app.route('/search_student')
def search_student():
    if 'user' not in session:
        flash("Please log in first.", "warning")
        return redirect(url_for('login'))

    query = request.args.get('query', '').lower()
    batch = request.args.get('batch', '')
    user_id = session['user']['id']

    # Initial query
    students_query = Student.query.filter_by(user_id=user_id)

    # Filter by name or roll if query exists
    if query:
        students_query = students_query.filter(
            db.or_(
                db.func.lower(Student.name).like(f'%{query}%'),
                Student.roll.like(f'%{query}%')
            )
        )

    # Filter by batch if selected
    if batch:
        students_query = students_query.filter_by(batch=batch)

    students = students_query.all()

    # Get all available batches for dropdown
    all_batches = db.session.query(Student.batch).filter_by(user_id=user_id).distinct().all()
    batches = sorted([b[0] for b in all_batches if b[0]])

    return render_template(
        'dashboard.html',
        students=students,
        user=session['user'],
        batches=batches,
        selected_batch=batch
    )

@app.route('/add_student', methods=['POST'])
def add_student():
    roll = request.form['roll']
    batch = request.form['batch']
    name = request.form['name']
    college = request.form['college']
    student_number = request.form.get('student_number', '')
    guardian_number = request.form.get('guardian_number', '')

    user_id = session['user']['id']

    existing = Student.query.filter_by(user_id=user_id, roll=roll).first()
    if existing:
        flash("⚠️ This roll already exists for your account!", "danger")
        return redirect(url_for('dashboard'))

    new_student = Student(
        roll=roll,
        batch=batch,
        name=name,
        college=college,
        student_number=student_number,
        guardian_number=guardian_number,
        user_id=user_id
    )
    db.session.add(new_student)
    db.session.commit()

    flash("✅ Student added successfully!", "success")
    return redirect(url_for('dashboard'))
@app.route('/edit_student/<roll>', methods=['GET', 'POST'])
def edit_student(roll):
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    student = Student.query.filter_by(user_id=user_id, roll=roll).first()

    if not student:
        flash("⚠️ You are not allowed to edit this student!", "danger")
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        student.batch = request.form['batch']
        student.name = request.form['name']
        student.college = request.form['college']
        student.student_number = request.form.get('student_number', '')
        student.guardian_number = request.form.get('guardian_number', '')

        db.session.commit()
        flash("✅ Student updated!", "success")
        return redirect(url_for('dashboard'))

    return render_template('edit_student.html', student=student, user=session['user'])
@app.route('/delete_student/<roll>')
def delete_student(roll):
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    student = Student.query.filter_by(user_id=user_id, roll=roll).first()

    if student:
        db.session.delete(student)
        db.session.commit()
        flash("🗑️ Student deleted", "info")

    return redirect(url_for('dashboard'))
@app.route('/get_next_roll')
def get_next_roll():
    if 'user' not in session:
        return jsonify({'next_roll': ''})

    user_id = session['user']['id']
    batch = request.args.get('batch', '')

    students = Student.query.filter_by(user_id=user_id, batch=batch).all()
    rolls = [int(s.roll) for s in students if s.roll.isdigit()]

    next_roll = str(max(rolls) + 1) if rolls else ''
    return jsonify({'next_roll': next_roll})
@app.route('/get_batch_by_roll_prefix')
def get_batch_by_roll_prefix():
    if 'user' not in session:
        return jsonify({'batch': ''})

    user_id = session['user']['id']
    prefix = request.args.get('prefix', '')

    student = Student.query.filter(
        Student.user_id == user_id,
        Student.roll.like(f'{prefix}%')
    ).first()

    return jsonify({'batch': student.batch if student else ''})
from flask import request, render_template, redirect, url_for, flash, session, jsonify
from models import db, Student, Payment, NewPayment

@app.route('/payment', methods=['GET', 'POST'])
def payment():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    user_students = Student.query.filter_by(user_id=user_id).all()
    students_dict = {student.roll: student for student in user_students}
    batches = sorted(set(s.batch for s in user_students if s.batch))
    selected_batch = request.args.get('batch', '')
    selected_status = request.args.get('status', '')

    roll = request.args.get('roll') or request.form.get('roll')
    student = students_dict.get(roll)

    payment_info = {
        "date": "",
        "memo_no": "",
        "receipt_no": "",
        "course": "",
        "total_payment": "",
        "previous_payment": 0,
        "new_payment": 0,
        "discount": 0,
        "due": 0,
        "reference": "",
        "status": "Due"
    }

    if roll and student:
        payment_record = Payment.query.filter_by(user_id=user_id, roll=roll).first()
        new_payment_entries = NewPayment.query.filter_by(user_id=user_id, roll=roll).all()

        if request.method == 'POST':
            date = request.form.get('date', '').strip()
            memo_no = request.form.get('memo_no', '').strip()
            session['last_memo_no'] = memo_no

            receipt_no_input = request.form.get('receipt_no', '').strip()
            total = int(request.form['total_payment'])
            previous = int(request.form.get('previous_payment', 0))
            new_payment_amount = int(request.form.get('new_payment', 0))
            discount = int(request.form.get('discount', 0))
            course = request.form.get('course', '').strip()
            if course == 'custom':
                course = request.form.get('custom_course', '').strip()


            reference = request.form.get('reference', '').strip()

            memo_payments = NewPayment.query.filter_by(user_id=user_id, memo_no=memo_no).all()

            if memo_payments:
                try:
                    max_receipt_no = max(int(p.receipt_no) for p in memo_payments if p.receipt_no.isdigit())
                except ValueError:
                    max_receipt_no = 0
                receipt_no = str(max_receipt_no + 1)
            else:
                receipt_no = receipt_no_input or "1"

            # ✅ Global Duplicate Check: Memo + Receipt unique per user
            duplicate_receipt_global = NewPayment.query.filter_by(
                user_id=user_id,
                memo_no=memo_no,
                receipt_no=receipt_no
            ).first()

            if duplicate_receipt_global:
                flash(f"❌ Receipt No '{receipt_no}' already exists under Memo No '{memo_no}' for this user.", "danger")
                return redirect(url_for('payment', roll=roll, batch=selected_batch, status=selected_status))

            updated_previous = previous + new_payment_amount
            due = total - (updated_previous + discount)
            status = "Paid" if due <= 0 else "Due"

            if not payment_record:
                payment_record = Payment(user_id=user_id, roll=roll, name=student.name, batch=student.batch)
                db.session.add(payment_record)

            payment_record.date = date
            payment_record.course = course
            payment_record.total_payment = total
            payment_record.previous_payment = updated_previous
            payment_record.discount = discount
            payment_record.due = due
            payment_record.reference = reference
            payment_record.status = status

            new_entry = NewPayment(
                user_id=user_id,
                roll=roll,
                new_payment=new_payment_amount,
                date=date,
                memo_no=memo_no,
                receipt_no=receipt_no,
                course=course
            )

            db.session.add(new_entry)
            db.session.commit()
            flash("✅ Payment updated", "success")
            return redirect(url_for('payment', roll=roll, batch=selected_batch, status=selected_status))

        else:
            memo_no_get = request.args.get('memo_no')
            memo_no = memo_no_get if memo_no_get else session.get('last_memo_no', '')

            memo_payments = NewPayment.query.filter_by(user_id=user_id, memo_no=memo_no).all()
            if memo_payments:
                try:
                    max_receipt_no = max(int(p.receipt_no) for p in memo_payments if p.receipt_no.isdigit())
                    next_receipt_no = str(max_receipt_no + 1)
                except ValueError:
                    next_receipt_no = ''
            else:
                next_receipt_no = ''

            payment_info = {
                "date": payment_record.date if payment_record else "",
                "memo_no": memo_no,
                "receipt_no": next_receipt_no,
                "course": payment_record.course if payment_record else "",
                "total_payment": payment_record.total_payment if payment_record else 0,
                "previous_payment": payment_record.previous_payment if payment_record else 0,
                "new_payment": 0,
                "due": payment_record.due if payment_record else 0,
                "discount": payment_record.discount if payment_record else 0,
                "reference": payment_record.reference if payment_record else "",
                "status": "Paid" if (payment_record and payment_record.due <= 0) else "Due"
            }

    filtered_students = [s for s in user_students if not selected_batch or s.batch == selected_batch]

    filtered_payments_query = Payment.query.filter_by(user_id=user_id)
    if selected_batch:
        filtered_payments_query = filtered_payments_query.filter(Payment.batch == selected_batch)
    if selected_status:
        filtered_payments_query = filtered_payments_query.filter(Payment.status == selected_status)

    filtered_payments = {
        p.roll: {
            "batch": p.batch,
            "name": p.name,
            "date": getattr(p, 'date', ''),
            "course": getattr(p, 'course', ''),
            "total_payment": p.total_payment,
            "previous_payment": p.previous_payment,
            "discount": p.discount,
            "due": p.due,
            "reference": p.reference,
            "status": p.status
        }
        for p in filtered_payments_query.all()
    }

    return render_template(
        'payment.html',
        students=filtered_students,
        payments=filtered_payments,
        batches=batches,
        selected_batch=selected_batch,
        selected_status=selected_status,
        student=student,
        payment=payment_info,
        roll=roll,
        filtered_payments=filtered_payments,
        user=session['user']
    )


@app.route('/download_memo_excel')
def download_memo_excel():
    if 'user' not in session:
        flash("Please login first", "warning")
        return redirect(url_for('login'))

    user_id = session['user']['id']
    memo_no = request.args.get('memo_no', '').strip()

    if not memo_no:
        flash("Memo No is required to download Excel.", "danger")
        return redirect(url_for('payment'))

    entries = (
        NewPayment.query
        .join(Student, NewPayment.roll == Student.roll)
        .filter(Student.user_id == user_id, NewPayment.memo_no == memo_no)
        .all()
    )

    if not entries:
        flash(f"No payments found for Memo No: {memo_no}", "info")
        return redirect(url_for('payment'))
    data = []
    for p in entries:
        student = p.student
        pay = Payment.query.filter_by(student_id=student.id).first()
        data.append({
            "Roll": student.roll,
            "Name": student.name,
            "Batch": student.batch,
            "Date": p.date,
            "Memo No": p.memo_no,
            "Receipt No": p.receipt_no,
            "Course": p.course,
            "New Payment": p.new_payment,
            "Total Payment": pay.total_payment if pay else 0,
            "Previous Payment": pay.previous_payment if pay else 0,
            "Discount": pay.discount if pay else 0,
            "Due": pay.due if pay else 0,
            "Status": pay.status if pay else "Due"
        })

    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=f'Memo_{memo_no}')
        ws = writer.sheets[f'Memo_{memo_no}']
        for col in ws.columns:
            length = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = length + 2
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

    output.seek(0)
    filename = f"memo_{memo_no}_payments.xlsx"
    return send_file(output, as_attachment=True,
                     download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

from collections import defaultdict

from collections import defaultdict
from openpyxl.styles import Font, Alignment

from openpyxl.styles import Font, Alignment

@app.route('/download_memo_list_excel')
def download_memo_list_excel():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    memo_filter = request.args.get('memo_no', '').strip()

    query = NewPayment.query.filter_by(user_id=user_id)
    if memo_filter:
        query = query.filter(NewPayment.memo_no == memo_filter)
    entries = query.order_by(NewPayment.memo_no, NewPayment.date, NewPayment.receipt_no).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Memo Receipts"

    headers = ["Date", "Memo No", "Receipt No", "Roll", "Name", "Course", "Amount"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    grand_total = 0
    current_memo = None
    memo_total = 0
    grand_total_so_far = 0

    for e in entries:
        student = e.student
        amount = e.new_payment or 0

        if current_memo and current_memo != e.memo_no:
            # ✅ Memo total
            ws.append(["", "", "", "", "", f"{current_memo} Total:", memo_total])
            total_row = ws.max_row
            ws[f"F{total_row}"].font = Font(bold=True)
            ws[f"G{total_row}"].font = Font(bold=True)

            # ✅ Grand total so far
            grand_total_so_far += memo_total
            ws.append(["", "", "", "", "", "Grand Total So Far:", grand_total_so_far])
            gts_row = ws.max_row
            ws[f"F{gts_row}"].font = Font(bold=True)
            ws[f"G{gts_row}"].font = Font(bold=True)

            # ✅ Blank row
            ws.append([])

            memo_total = 0

        ws.append([
            e.date or '',
            e.memo_no or '',
            e.receipt_no or '',
            student.roll if student else e.roll,
            student.name if student else '',
            e.course or '',
            amount
        ])

        current_memo = e.memo_no
        memo_total += amount
        grand_total += amount

    # ✅ Final memo total and grand total so far
    if current_memo:
        ws.append(["", "", "", "", "", f"{current_memo} Total:", memo_total])
        total_row = ws.max_row
        ws[f"F{total_row}"].font = Font(bold=True)
        ws[f"G{total_row}"].font = Font(bold=True)

        grand_total_so_far += memo_total
        ws.append(["", "", "", "", "", "Grand Total So Far:", grand_total_so_far])
        gts_row = ws.max_row
        ws[f"F{gts_row}"].font = Font(bold=True)
        ws[f"G{gts_row}"].font = Font(bold=True)

        # ✅ Blank row
        ws.append([])

    # ✅ Final Grand Total
    ws.append(["", "", "", "", "", "Grand Total:", grand_total])
    final_row = ws.max_row
    ws[f"F{final_row}"].font = Font(bold=True)
    ws[f"G{final_row}"].font = Font(bold=True)

    # ✅ Autosize columns
    for col_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"memo_receipt_list_{memo_filter or 'all'}.xlsx"
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


from math import ceil
from flask import (
    Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
)
from models import db, User, Student, Payment, NewPayment
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
def is_duplicate_receipt(user_id, roll, memo_no, receipt_no):
    return NewPayment.query.filter_by(user_id=user_id, roll=roll, memo_no=memo_no, receipt_no=receipt_no).first() is not None



@app.route('/get_next_receipt_no')
def get_next_receipt_no():
    if 'user' not in session:
        return jsonify(status="error", message="Unauthorized"), 401

    memo_no = request.args.get('memo_no', '').strip()
    if not memo_no:
        return jsonify(status="error", message="Missing memo_no"), 400

    user_id = session['user']['id']

    # Memo No অনুযায়ী ওই ইউজারের সব পেমেন্ট বের করো
    existing = NewPayment.query.filter_by(user_id=user_id, memo_no=memo_no).all()

    if existing:
        try:
            max_receipt = max(int(e.receipt_no) for e in existing if e.receipt_no.isdigit())
        except ValueError:
            max_receipt = 0
        next_no = str(max_receipt + 1)
    else:
        next_no = "1"

    return jsonify(status="success", receipt_no=next_no)




@app.route('/memo_list')
def memo_list():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    search_memo = request.args.get('memo_no', '').strip()
    search_roll = request.args.get('roll', '').strip()
    search_name = request.args.get('name', '').strip()
    receipt_no = request.args.get('receipt_no', '')
    page = int(request.args.get('page', 1))
    per_page = 50

    # ✅ Filter directly by NewPayment.user_id
    query = NewPayment.query.filter(NewPayment.user_id == user_id)

    # Optional search filters
    if search_memo:
        query = query.filter(NewPayment.memo_no.contains(search_memo))
    if search_roll:
        query = query.filter(NewPayment.roll == search_roll)
    if search_name:
        # Join Student only if name search is required
        query = query.join(Student, NewPayment.roll == Student.roll)\
                     .filter(Student.user_id == user_id)\
                     .filter(Student.name.ilike(f"%{search_name}%"))
    if receipt_no:
        query = query.filter(NewPayment.receipt_no.contains(receipt_no))

    query = query.order_by(NewPayment.memo_no, NewPayment.receipt_no)
    total = query.count()
    total_pages = ceil(total / per_page)
    entries = query.offset((page - 1) * per_page).limit(per_page).all()

    memos = [{
        'memo_no': e.memo_no,
        'receipt_no': e.receipt_no,
        'roll': e.roll,
        'name': e.student.name if e.student else '',
        'date': e.date,
        'amount': e.new_payment,
        'course': e.course
    } for e in entries]

    return render_template(
        'memo_list.html',
        memos=memos,
        page=page,
        total_pages=total_pages,
        search_memo_no=search_memo,
        search_roll=search_roll,
        search_name=search_name
    )

from flask import send_file
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import cm
from datetime import datetime
import io

@app.route('/export_memo_receipt_pdf')
def export_memo_receipt_pdf():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    memo_filter = request.args.get('memo_no', '').strip()

    query = NewPayment.query.filter_by(user_id=user_id)
    if memo_filter:
        query = query.filter(NewPayment.memo_no == memo_filter)

    entries = query.order_by(NewPayment.memo_no, NewPayment.receipt_no).all()
    if not entries:
        flash(f"No memo entries found for Memo No: {memo_filter}", "info")
        return redirect(url_for('memo_list'))

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)
    margin = 1.5 * cm
    row_height = 20

    headers = ["Date", "Memo No", "Receipt No", "Roll", "Name", "Course", "Amount"]
    col_widths = [3*cm, 2*cm, 3*cm, 2.5*cm, 8*cm, 6*cm, 3*cm]
    x_positions = [margin]
    for w in col_widths[:-1]:
        x_positions.append(x_positions[-1] + w)
    end_x = x_positions[-1] + col_widths[-1]

    y = height - margin
    row_count = 0
    page_total = 0
    grand_total_so_far = 0
    current_memo = None
    max_rows_per_page = 40

    def draw_table_header(memo_no):
        nonlocal y
        y = height - margin
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawCentredString(width / 2, y, "Memo & Receipt Payment Report")
        y -= row_height
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawCentredString(width / 2, y, f"Memo No: {memo_no}")
        y -= row_height * 1.5
        pdf.setFont("Helvetica-Bold", 10)
        for i, header in enumerate(headers):
            pdf.drawCentredString(x_positions[i] + col_widths[i]/2, y + 5, header)
        pdf.line(margin, y + row_height, end_x, y + row_height)
        pdf.line(margin, y, end_x, y)
        for x in x_positions + [end_x]:
            pdf.line(x, y, x, y + row_height)
        y -= row_height

    def draw_footer(page_total, grand_total):
        nonlocal y
        y -= row_height / 2
        pdf.line(margin, y, end_x, y)
        y -= row_height
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawRightString(end_x - 2, y + 5, f"Page Total: {page_total:.2f}")
        pdf.drawRightString(end_x - 2, y - 10, f"Grand Total So Far: {grand_total:.2f}")
        y -= row_height * 1.5

    for entry in entries:
        if current_memo != entry.memo_no or row_count >= max_rows_per_page or y < margin + 3*row_height:
            if current_memo is not None:
                grand_total_so_far += page_total
                draw_footer(page_total, grand_total_so_far)
                pdf.showPage()

            current_memo = entry.memo_no
            page_total = 0
            row_count = 0
            draw_table_header(current_memo)

        student = entry.student
        raw_date = entry.date
        try:
            formatted_date = datetime.strptime(raw_date, "%Y-%m-%d").strftime("%d/%m/%Y")
        except:
            formatted_date = raw_date

        row_data = [
            formatted_date, entry.memo_no, entry.receipt_no,
            student.roll if student else entry.roll,
            student.name if student else '',
            entry.course,
            f"{float(entry.new_payment):.2f}"
        ]

        pdf.setFont("Helvetica", 10)
        for i, val in enumerate(row_data):
            pdf.drawCentredString(x_positions[i] + col_widths[i]/2, y + 5, str(val))

        pdf.line(margin, y, end_x, y)
        for x in x_positions + [end_x]:
            pdf.line(x, y, x, y + row_height)

        y -= row_height
        page_total += float(entry.new_payment or 0)
        row_count += 1

    # ✅ Last page totals
    grand_total_so_far += page_total
    draw_footer(page_total, grand_total_so_far)

    pdf.save()
    buffer.seek(0)

    return send_file(buffer, as_attachment=True,
                     download_name="memo_receipt_report.pdf", mimetype='application/pdf')




@app.route('/memo/edit/<memo_no>/<receipt_no>', methods=['GET', 'POST'])
def edit_memo_receipt(memo_no, receipt_no):
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    # student_exists = Student.query.filter_by(user_id=user_id).exists()  # not used; can be removed
    entry = NewPayment.query.join(Student, NewPayment.roll == Student.roll).filter(
        Student.user_id == user_id,
        NewPayment.memo_no == memo_no,
        NewPayment.receipt_no == receipt_no
    ).first()

    if not entry:
        flash("❌ Memo entry not found", "danger")
        return redirect(url_for('memo_list'))

    if request.method == 'POST':
        entry.date = request.form.get('date', '').strip()
        entry.memo_no = request.form.get('memo_no', '').strip()
        entry.receipt_no = request.form.get('receipt_no', '').strip()
        entry.course = request.form.get('course', '').strip()
        entry.new_payment = float(request.form.get('new_payment', 0) or 0)

        db.session.commit()
        flash("✅ Memo updated successfully", "success")
        return redirect(url_for('memo_list'))

    return render_template('edit_memo.html', entry=entry)


@app.route('/memo/delete/<memo_no>/<receipt_no>', methods=['POST'])
def delete_memo_receipt(memo_no, receipt_no):
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    entry = NewPayment.query.filter_by(
        user_id=user_id,
        memo_no=memo_no,
        receipt_no=receipt_no
    ).first()

    if not entry:
        flash("❌ Memo receipt not found or already deleted.", "warning")
    else:
        db.session.delete(entry)
        db.session.commit()
        flash("🗑️ Memo receipt deleted successfully.", "success")

    return redirect(url_for('memo_list'))


@app.route('/export_excel_custom')
def export_excel_custom():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    batch = request.args.get('batch')
    download = request.args.get('download') == '1'

    query = Student.query.filter_by(user_id=user_id)
    if batch:
        query = query.filter_by(batch=batch)
    students = query.all()

    if not students:
        flash("⚠️ No student data to export", "warning")
        return redirect(url_for('dashboard', batch=batch))

    df = pd.DataFrame([{
        "roll": s.roll,
        "batch": s.batch,
        "name": s.name,
        "college": s.college,
        "student_number": s.student_number,
        "guardian_number": s.guardian_number
    } for s in students])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Students')
    output.seek(0)

    filename = f"students_{batch or 'all'}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)
@app.route('/view_excel')
def view_excel():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    batch = request.args.get('batch')
    query = Student.query.filter_by(user_id=user_id)
    if batch and batch.lower() != 'all':
        query = query.filter_by(batch=batch)
    students = query.all()

    return render_template('view_excel.html', students=students, batch=batch)
@app.route('/import_excel', methods=['POST'])
def import_excel():
    if 'user' not in session:
        flash("⚠️ প্রথমে লগইন করুন।", "warning")
        return redirect(url_for('login'))

    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        flash("⚠️ একটি বৈধ .xlsx ফাইল আপলোড করুন!", "danger")
        return redirect(url_for('dashboard'))

    try:
        df = pd.read_excel(file)
    except Exception as e:
        flash("⚠️ ফাইলটি পড়তে সমস্যা হয়েছে!", "danger")
        return redirect(url_for('dashboard'))

    required = {'name', 'roll', 'batch', 'student_number', 'guardian_number'}
    if not required.issubset(df.columns):
        flash("⚠️ Excel-এ অবশ্যই name, roll, batch, student_number, guardian_number কলাম থাকতে হবে!", "danger")
        return redirect(url_for('dashboard'))

    user_id = session['user']['id']
    existing_rolls = {s.roll for s in Student.query.filter_by(user_id=user_id).all()}
    added = 0
    skipped = 0

    for index, row in df.iterrows():
        roll = str(row.get('roll', '')).strip()
        if not roll or roll in existing_rolls:
            skipped += 1
            continue

        # ডেটা সেফভাবে নেয়া
        name = str(row.get('name', '')).strip()
        batch = str(row.get('batch', '')).strip()
        college = str(row.get('college', '')).strip()
        student_number = str(row.get('student_number', '')).strip()
        guardian_number = str(row.get('guardian_number', '')).strip()

        # নতুন স্টুডেন্ট তৈরি
        new = Student(
            roll=roll,
            name=name,
            batch=batch,
            college=college,
            student_number=student_number,
            guardian_number=guardian_number,
            user_id=user_id
        )
        db.session.add(new)
        existing_rolls.add(roll)
        added += 1

    db.session.commit()

    flash(f"✅ {added} জন নতুন শিক্ষার্থী সফলভাবে ইম্পোর্ট হয়েছে!", "success")
    if skipped > 0:
        flash(f"⚠️ {skipped}টি রো স্কিপ করা হয়েছে (রোল ডুপ্লিকেট বা ডেটা মিসিং)।", "warning")
    return redirect(url_for('dashboard'))

@app.route('/import_payment_excel', methods=['POST'])
def import_payment_excel():
    if 'user' not in session:
        flash("⚠️ Please log in first.", "warning")
        return redirect(url_for('login'))

    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        flash("⚠️ Please upload a valid .xlsx file!", "danger")
        return redirect(url_for('dashboard'))

    try:
        df = pd.read_excel(file)
    except Exception as e:
        flash("⚠️ Failed to read the file!", "danger")
        return redirect(url_for('dashboard'))

    required = {'roll', 'total_payment', 'previous_payment', 'discount', 'reference'}
    if not required.issubset(df.columns):
        flash("⚠️ Excel file must contain the following columns: roll, total_payment, previous_payment, discount, reference!", "danger")
        return redirect(url_for('dashboard'))


    # সেফলি সংখ্যা রূপান্তরের জন্য সহায়ক ফাংশন
    def safe_int(val):
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return 0

    user_id = session['user']['id']
    updated = 0
    skipped = 0

    for _, row in df.iterrows():
        roll = str(row.get('roll', '')).strip()
        if not roll:
            skipped += 1
            continue

        student = Student.query.filter_by(user_id=user_id, roll=roll).first()
        if student:
            total = safe_int(row.get('total_payment'))
            previous = safe_int(row.get('previous_payment'))
            discount = safe_int(row.get('discount'))
            reference = str(row.get('reference', '')).strip()

            due = total - (previous + discount)
            status = "Paid" if due <= 0 else "Due"

            # ✅ পেমেন্ট রেকর্ড খোঁজা এবং তৈরি/আপডেট করা
            payment = Payment.query.filter_by(roll=student.roll, user_id=user_id).first()
            if not payment:
                payment = Payment(
                    roll=student.roll,
                    name=student.name,
                    batch=student.batch,
                    user_id=user_id
                )
                db.session.add(payment)

            # ফিল্ড আপডেট
            payment.total_payment = total
            payment.previous_payment = previous
            payment.discount = discount
            payment.due = due
            payment.reference = reference
            payment.status = status

            updated += 1
        else:
            skipped += 1

    db.session.commit()

    flash(f"✅ Payment data imported for {updated} student(s)!", "success")
    if skipped > 0:
        flash(f"⚠️ {skipped} row(s) were skipped (possibly due to missing roll or invalid data).", "warning")
    return redirect(url_for('dashboard'))


from io import BytesIO
import pandas as pd
from flask import flash, redirect, url_for, session, request, render_template, send_file
from fpdf import FPDF
from models import Payment, Student

@app.route('/view_payment_excel')
def view_payment_excel():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    batch = request.args.get('batch', '')
    status = request.args.get('status', '')

    query = Payment.query.join(Student, Payment.roll == Student.roll).filter(Student.user_id == user_id)
    if batch:
        query = query.filter(Student.batch == batch)
    if status:
        query = query.filter(Payment.status == status)

    data = [{
        'roll': p.student.roll,
        'name': p.student.name,
        'batch': p.student.batch,
        'total_payment': p.total_payment,
        'previous_payment': p.previous_payment,
        'discount': p.discount,
        'reference': p.reference,
        'due': p.due,
        'status': p.status
    } for p in query.all()]

    return render_template('view_payment_excel.html', payments=data, batch=batch, status=status)


@app.route('/export_payment_excel')
def export_payment_excel():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    batch = request.args.get('batch', '')
    status = request.args.get('status', '')
    download = request.args.get('download', '0') == '1'

    query = Payment.query.join(Student, Payment.roll == Student.roll).filter(Student.user_id == user_id)
    if batch:
        query = query.filter(Student.batch == batch)
    if status:
        query = query.filter(Payment.status == status)

    rows = query.all()
    if not rows:
        flash("⚠️ No payment data found for export!", "warning")
        return redirect(url_for('dashboard'))

    df = pd.DataFrame([{
        'Roll': p.student.roll,
        'Name': p.student.name,
        'Batch': p.student.batch,
        'Total Payment': p.total_payment,
        'Previous Payment': p.previous_payment,
        'Discount': p.discount,
        'Reference': p.reference,
        'Due': p.due,
        'Status': p.status
    } for p in rows])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Payments')
    output.seek(0)

    filename = f"payment_report_{batch or 'all'}_{status or 'all'}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=download,
                     download_name=filename)


@app.route('/export_payment_pdf')
def export_payment_pdf():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']
    batch = request.args.get('batch', '')
    status_filter = request.args.get('status', '')
    download = request.args.get('download', '0') == '1'

    query = Payment.query.join(Student, Payment.roll == Student.roll).filter(Student.user_id == user_id)
    if batch:
        query = query.filter(Student.batch == batch)
    if status_filter:
        query = query.filter(Payment.status == status_filter)

    rows = query.all()
    if not rows:
        flash("⚠️ No payment data found!", "warning")
        return redirect(url_for('dashboard'))

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font('Arial', 'B', 30)  # 'B' for bold (optional)
    pdf.cell(190, 15, 'RETINA', ln=True, align='C')
    pdf.set_font("Arial", size=15)
    header = f"Payment Report - Batch: {batch or 'All'} | Status: {status_filter or 'All'}"
    pdf.cell(190, 10, txt=header, ln=True, align='C')
    pdf.ln(5)

    headers = ["Roll", "Name", "Batch", "Total", "Prev", "Discount", "Ref", "Due", "Status"]
    col_widths = [15, 50, 15, 20, 20, 20, 20, 15, 15]

    pdf.set_font("Arial", 'B', 10)
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 8, txt=h, border=1, align='C')
    pdf.ln()

    pdf.set_font("Arial", size=10)
    for p in rows:
        row = [
            p.student.roll,
            p.student.name,
            p.student.batch,
            str(p.total_payment),
            str(p.previous_payment),
            str(p.discount),
            str(p.reference),
            str(p.due),
            p.status
        ]
        for i, val in enumerate(row):
            align = 'L' if i == 1 else 'C'
            pdf.cell(col_widths[i], 8, txt=val[:30] if i == 1 else val[:12], border=1, align=align)
        pdf.ln()

    pdf_io = BytesIO(pdf.output(dest='S').encode('latin1'))
    return send_file(pdf_io,
                     mimetype='application/pdf',
                     as_attachment=download,
                     download_name=f"payment_report_{batch or 'all'}_{status_filter or 'all'}.pdf")

@app.route('/delete_payment/<roll>')
def delete_payment(roll):
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = session['user']['id']

    # Payment টা roll আর user_id দিয়ে খুঁজবে
    payment = Payment.query.filter_by(user_id=user_id, roll=roll).first()
    if payment:
        db.session.delete(payment)
        db.session.commit()
        flash(f"🗑️ Payment deleted for Roll {roll}", "success")
    else:
        flash(f"⚠️ No payment found for Roll {roll}", "warning")

    return redirect(url_for('payment'))

@app.route('/register', methods=['GET','POST'])
def register():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email'].strip().lower()
        phone = request.form['phone'].strip()
        raw_password = request.form['password'].strip()
        if not (email or phone):
            return render_template('register.html', error="Email or Phone দিতে হবে")
        if User.query.filter((User.email==email)|(User.phone==phone)).first():
            return render_template('register.html', error="Email বা Phone already used")
        new = User(
            uuid=str(uuid.uuid4()), name=name,
            email=email or None, phone=phone or None,
            password=generate_password_hash(raw_password)
        )
        db.session.add(new)
        db.session.commit()
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        login_id = request.form['login'].strip().lower()
        pwd = request.form['password'].strip()
        user = User.query.filter((User.email==login_id)|(User.phone==login_id)).first()
        if user and check_password_hash(user.password, pwd):
            session['user'] = {'id': user.id, 'is_admin': user.is_admin, 'name': user.name}
            return redirect(url_for('dashboard'))
        return render_template('login.html', error="Invalid credentials")
    return render_template('login.html')
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/forgot', methods=['GET','POST'])
def forgot():
    if request.method == 'POST':
        login_id = request.form['login'].strip().lower()
        user = User.query.filter((User.email==login_id)|(User.phone==login_id)).first()
        if user:
            return redirect(url_for('reset_password', user_uuid=user.uuid))
        return render_template('forgot_password.html', error="User not found")
    return render_template('forgot_password.html')

@app.route('/reset/<user_uuid>', methods=['GET','POST'])
def reset_password(user_uuid):
    user = User.query.filter_by(uuid=user_uuid).first()
    if not user:
        return redirect(url_for('forgot'))
    if request.method == 'POST':
        user.password = generate_password_hash(request.form['password'].strip())
        db.session.commit()
        return redirect(url_for('login'))
    return render_template('reset_password.html')
@app.route('/admin')
def admin():
    if not session.get('user') or not session['user']['is_admin']:
        return redirect('/')
    users = User.query.all()
    return render_template('admin.html', users=users)

@app.route('/admin/delete/<string:user_id>')
def admin_delete(user_id):
    if not session.get('user', {}).get('is_admin'):
        return redirect('/')

    user = User.query.get(user_id)
    if user:
        # প্রথমে user এর payment গুলো ডিলেট করো
        Payment.query.filter_by(user_id=user_id).delete()
        Student.query.filter_by(user_id=user_id).delete()
        
        # তারপর ইউজার ডিলেট করো
        db.session.delete(user)
        db.session.commit()
    return redirect(url_for('admin'))


@app.route('/admin/promote/<string:user_id>')
def admin_promote(user_id):
    user = User.query.get(user_id)
    if session.get('user',{}).get('is_admin') and user:
        user.is_admin = True
        db.session.commit()
    return redirect(url_for('admin'))

@app.route('/admin/demote/<string:user_id>')
def admin_demote(user_id):
    user = User.query.get(user_id)
    if session.get('user',{}).get('is_admin') and user:
        user.is_admin = False
        db.session.commit()
    return redirect(url_for('admin'))

@app.route('/admin/edit/<string:user_id>', methods=['GET','POST'])
def admin_edit(user_id):
    if not session.get('user',{}).get('is_admin'):
        return redirect('/')
    user = User.query.get(user_id)
    if not user:
        return redirect(url_for('admin'))
    if request.method == 'POST':
        user.name = request.form['name'].strip() or user.name
        user.email = request.form['email'].strip().lower() or user.email
        user.phone = request.form['phone'].strip() or user.phone
        db.session.commit()
        return redirect(url_for('admin'))
    return render_template('admin_edit.html', user=user)

@app.route('/view_students')
def view_students():
    if not session.get('user'):
        return redirect(url_for('login'))
    user_id = session['user']['id']
    batch = request.args.get('batch','')
    query = Student.query.filter_by(user_id=user_id)
    if batch:
        query = query.filter_by(batch=batch)
    students = query.all()
    batches = sorted({s.batch for s in students if s.batch})
    return render_template('view_students.html', students=students, batches=batches, selected_batch=batch)

@app.route('/view_payments')
def view_payments():
    if not session.get('user'):
        return redirect(url_for('login'))
    user_id = session['user']['id']
    batch = request.args.get('batch', '')
    
    query = Payment.query.filter_by(user_id=user_id)
    if batch:
        query = query.filter(Payment.batch == batch)
    
    payments = query.all()
    batches = sorted({p.batch for p in payments if p.batch})
    
    return render_template('view_payments.html', payments=payments, batches=batches, selected_batch=batch)


@app.route('/delete_all_students', methods=['POST'])
def delete_all_students():
    if 'user' not in session:
        return redirect(url_for('login'))
    user_id = session['user']['id']
    users = User.query.get(user_id)
    if not check_password_hash(users.password, request.form.get('password')):
        flash("❌ Incorrect password!", "danger")
        return redirect(url_for('dashboard'))
    Student.query.filter_by(user_id=user_id).delete()
    db.session.commit()
    flash("✅ All students deleted successfully!", "success")
    return redirect(url_for('dashboard'))


@app.route('/delete_all_payments', methods=['POST'])
def delete_all_payments():
    if 'user' not in session:
        return redirect(url_for('login'))
    user_id = session['user']['id']
    users = User.query.get(user_id)
    if not check_password_hash(users.password, request.form.get('password')):
        flash("❌ Incorrect password!", "danger")
        return redirect(url_for('payment'))

    # student_id নাই, তাই সরাসরি user_id দিয়ে ফিল্টার করবো
    Payment.query.filter_by(user_id=user_id).delete(synchronize_session=False)
    db.session.commit()
    flash("✅ All payments deleted successfully!", "success")
    return redirect(url_for('payment'))




from flask import send_file, session, redirect, url_for, request
from io import BytesIO
from fpdf import FPDF
from models import Student  # SQLAlchemy Model

@app.route('/export_pdf_custom')
def export_pdf_custom():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_id = str(session['user']['id'])
    batch = request.args.get('batch')
    download = request.args.get('download') == '1'

    students_query = Student.query.filter_by(user_id=user_id)
    if batch and batch.lower() != 'all':
        students_query = students_query.filter_by(batch=batch)
    students = students_query.all()
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 30)  # 'B' for bold (optional)
    pdf.cell(190, 15, 'RETINA', ln=True, align='C')
    pdf.set_font("Arial", size=15)
    pdf.cell(190, 10, txt="Student List", ln=True, align='C')
    pdf.ln(5)

    col_widths = [10, 15, 15, 60, 30, 30, 30]
    row_height = 8
    headers = ["SL", "Roll", "Batch", "Name", "College", "Student No", "Guardian No"]

    def add_table_header():
        pdf.set_font("Arial", 'B', 10)
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], row_height, header, 1, 0, 'C')
        pdf.ln()

    add_table_header()
    pdf.set_font("Arial", '', 10)

    for idx, s in enumerate(students, start=1):
        if pdf.get_y() > 270:
            pdf.add_page()
            add_table_header()
            pdf.set_font("Arial", '', 10)

        data = [
            str(idx),
            s.roll or "",
            s.batch or "",
            s.name or "",
            s.college or "",
            s.student_number or "",
            s.guardian_number or ""
        ]

        for i, txt in enumerate(data):
            y_before = pdf.get_y()
            x_before = pdf.get_x()

            pdf.cell(col_widths[i], row_height, '', 1)
            v_offset = (row_height - 5) / 2
            pdf.set_xy(x_before, y_before + v_offset)
            pdf.cell(col_widths[i], 5, txt, 0, 0, 'C')
            pdf.set_xy(x_before + col_widths[i], y_before)

        pdf.ln(row_height)

    pdf_data = pdf.output(dest='S').encode('latin1')
    pdf_output = BytesIO(pdf_data)

    return send_file(
        pdf_output,
        mimetype='application/pdf',
        as_attachment=download,
        download_name=f"students_{batch or 'all'}.pdf"
    )



if __name__ == '__main__':
    app.run(debug=True)
